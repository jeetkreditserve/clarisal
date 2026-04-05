from __future__ import annotations

import csv
import io
from abc import ABC, abstractmethod
from typing import Any

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class BaseReport(ABC):
    @property
    @abstractmethod
    def title(self) -> str:
        raise NotImplementedError

    @property
    @abstractmethod
    def columns(self) -> list[str]:
        raise NotImplementedError

    @abstractmethod
    def generate_rows(self) -> list[dict[str, Any]]:
        raise NotImplementedError

    def to_json(self) -> dict[str, Any]:
        return {
            'title': self.title,
            'columns': self.columns,
            'rows': self.generate_rows(),
        }

    def to_csv_response(self) -> HttpResponse:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.title.lower().replace(" ", "-")}.csv"'
        writer = csv.DictWriter(response, fieldnames=self.columns, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(self.generate_rows())
        return response

    def to_excel_response(self) -> HttpResponse:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.title[:31]
        worksheet.append(self.columns)

        for row in self.generate_rows():
            worksheet.append([row.get(column, '') for column in self.columns])

        for index, column in enumerate(self.columns, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = max(len(column) + 2, 14)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.title.lower().replace(" ", "-")}.xlsx"'
        return response

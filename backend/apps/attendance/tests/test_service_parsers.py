from datetime import date, datetime, time
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook

from apps.attendance.models import (
    AttendanceImportJob,
    AttendanceImportMode,
    AttendanceImportRow,
    AttendanceImportRowStatus,
    AttendanceImportStatus,
)
from apps.attendance.services import (
    _ensure_xlsx,
    _load_sheet_rows,
    _parse_date,
    _parse_datetime_value,
    _parse_time_value,
    _resolve_employee,
    build_attendance_sheet_sample,
    build_normalized_attendance_workbook,
    build_punch_sheet_sample,
)
from apps.timeoff.tests.test_services import _create_employee, _create_organisation


def _build_uploaded_workbook(name, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return SimpleUploadedFile(
        name,
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@pytest.mark.parametrize(
    ('raw_value', 'expected'),
    [
        (date(2026, 4, 1), date(2026, 4, 1)),
        (datetime(2026, 4, 1, 9, 30), date(2026, 4, 1)),
        ('2026-04-01', date(2026, 4, 1)),
        ('01-04-2026', date(2026, 4, 1)),
        ('01/04/2026', date(2026, 4, 1)),
        ('2026/04/01', date(2026, 4, 1)),
    ],
)
def test_parse_date_accepts_supported_formats(raw_value, expected):
    assert _parse_date(raw_value) == expected


def test_parse_date_rejects_invalid_value():
    with pytest.raises(ValueError, match='Date must be'):
        _parse_date('2026.04.01')


@pytest.mark.parametrize(
    ('raw_value', 'expected'),
    [
        (time(9, 5), time(9, 5)),
        (datetime(2026, 4, 1, 9, 5), time(9, 5)),
        ('09:05', time(9, 5)),
        ('09:05:10', time(9, 5, 10)),
        ('09:05 AM', time(9, 5)),
        ('2026-04-01 09:05', time(9, 5)),
        ('2026-04-01T09:05:10', time(9, 5, 10)),
    ],
)
def test_parse_time_value_accepts_supported_formats(raw_value, expected):
    assert _parse_time_value(raw_value, label='Check in') == expected


def test_parse_datetime_value_returns_aware_datetime_and_rejects_invalid():
    parsed = _parse_datetime_value('2026-04-01T09:05:10', label='Punch time')

    assert parsed.isoformat().startswith('2026-04-01T09:05:10')
    assert parsed.tzinfo is not None

    with pytest.raises(ValueError, match='Punch time must be a valid ISO datetime'):
        _parse_datetime_value('04/01/2026 09:05', label='Punch time')


@pytest.mark.django_db
def test_load_sheet_rows_validates_headers_and_skips_blank_rows():
    workbook = _build_uploaded_workbook(
        'attendance.xlsx',
        ['employee_code', 'date', 'check_in', 'check_out'],
        [
            ['EMP100', '2026-04-01', '09:05', '18:10'],
            [None, None, None, None],
            ['EMP101', '2026-04-02', '09:15', '18:15'],
        ],
    )

    rows = _load_sheet_rows(workbook, ['employee_code', 'date', 'check_in', 'check_out'])

    assert len(rows) == 2
    assert rows[0]['employee_code'] == 'EMP100'
    assert rows[0]['row_number'] == 2
    assert rows[1]['row_number'] == 4

    wrong_header_workbook = _build_uploaded_workbook(
        'bad.xlsx',
        ['employee', 'date', 'check_in', 'check_out'],
        [['EMP100', '2026-04-01', '09:05', '18:10']],
    )
    with pytest.raises(ValueError, match='must contain these columns'):
        _load_sheet_rows(wrong_header_workbook, ['employee_code', 'date', 'check_in', 'check_out'])


@pytest.mark.django_db
def test_resolve_employee_normalizes_code_and_requires_active_employee():
    organisation = _create_organisation('Attendance Parser Org')
    employee = _create_employee(organisation, email='attendance-parser@test.com')

    normalized_code, resolved_employee = _resolve_employee(organisation, employee.employee_code.lower())

    assert normalized_code == employee.employee_code
    assert resolved_employee == employee

    with pytest.raises(ValueError, match='Employee code is required'):
        _resolve_employee(organisation, '')

    employee.status = 'TERMINATED'
    employee.save(update_fields=['status', 'modified_at'])
    with pytest.raises(ValueError, match='was not found'):
        _resolve_employee(organisation, employee.employee_code)


@pytest.mark.django_db
def test_sample_and_normalized_workbooks_include_expected_sheets():
    organisation = _create_organisation('Attendance Workbook Org')
    employee = _create_employee(organisation, email='attendance-workbook@test.com')
    job = AttendanceImportJob.objects.create(
        organisation=organisation,
        mode=AttendanceImportMode.PUNCH_SHEET,
        status=AttendanceImportStatus.READY_FOR_REVIEW,
        original_filename='punches.xlsx',
    )
    AttendanceImportRow.objects.create(
        job=job,
        row_number=2,
        employee_code=employee.employee_code,
        employee=employee,
        attendance_date=date(2026, 4, 1),
        check_in_at=_parse_datetime_value('2026-04-01T09:05:00', label='Check in'),
        check_out_at=_parse_datetime_value('2026-04-01T18:10:00', label='Check out'),
        raw_punch_times=['2026-04-01T09:05:00', '2026-04-01T18:10:00'],
        status=AttendanceImportRowStatus.VALID,
    )
    AttendanceImportRow.objects.create(
        job=job,
        row_number=3,
        employee_code=employee.employee_code,
        employee=employee,
        attendance_date=date(2026, 4, 2),
        check_in_at=_parse_datetime_value('2026-04-02T09:10:00', label='Check in'),
        raw_punch_times=['2026-04-02T09:10:00'],
        status=AttendanceImportRowStatus.INCOMPLETE,
        error_message='Missing checkout punch.',
    )
    AttendanceImportRow.objects.create(
        job=job,
        row_number=4,
        employee_code='UNKNOWN',
        status=AttendanceImportRowStatus.ERROR,
        error_message='Employee not found.',
    )

    attendance_sample = load_workbook(BytesIO(build_attendance_sheet_sample()))
    punch_sample = load_workbook(BytesIO(build_punch_sheet_sample()))
    normalized = load_workbook(BytesIO(build_normalized_attendance_workbook(job)))

    assert attendance_sample.active['A1'].value == 'employee_code'
    assert 'Instructions' in attendance_sample.sheetnames
    assert punch_sample.active['C1'].value == 'punch_time'
    assert 'Review Notes' in normalized.sheetnames
    assert 'Validation Errors' in normalized.sheetnames
    assert normalized['Attendance Import']['A2'].value == employee.employee_code
    assert normalized['Validation Errors']['B2'].value == 'UNKNOWN'


def test_ensure_xlsx_rejects_non_excel_files():
    _ensure_xlsx(SimpleUploadedFile('attendance.xlsx', b'data'))

    with pytest.raises(ValueError, match='Only Excel .xlsx files are supported'):
        _ensure_xlsx(SimpleUploadedFile('attendance.csv', b'data'))

from datetime import date
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import AccountType, User, UserRole
from apps.approvals.models import (
    ApprovalApproverType,
    ApprovalRequestKind,
    ApprovalStage,
    ApprovalStageApprover,
    ApprovalWorkflow,
    ApprovalWorkflowRule,
)
from apps.approvals.services import approve_action
from apps.attendance.models import (
    AttendanceDay,
    AttendanceImportStatus,
    AttendancePunch,
    AttendanceRecord,
    AttendanceRegularizationRequest,
    AttendanceRegularizationStatus,
    AttendanceSourceConfig,
)
from apps.employees.models import Employee, EmployeeStatus
from apps.organisations.models import (
    Organisation,
    OrganisationAccessState,
    OrganisationBillingStatus,
    OrganisationMembership,
    OrganisationMembershipStatus,
    OrganisationStatus,
)


def build_xlsx(headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@pytest.fixture
def attendance_setup(db):
    organisation = Organisation.objects.create(
        name='Acme Attendance',
        status=OrganisationStatus.ACTIVE,
        billing_status=OrganisationBillingStatus.PAID,
        access_state=OrganisationAccessState.ACTIVE,
    )
    org_admin_user = User.objects.create_user(
        email='attendance-admin@test.com',
        password='pass123!',
        account_type=AccountType.WORKFORCE,
        role=UserRole.ORG_ADMIN,
        organisation=organisation,
        is_active=True,
    )
    OrganisationMembership.objects.create(
        user=org_admin_user,
        organisation=organisation,
        is_org_admin=True,
        status=OrganisationMembershipStatus.ACTIVE,
    )
    organisation.primary_admin_user = org_admin_user
    organisation.save(update_fields=['primary_admin_user', 'modified_at'])
    employee_user = User.objects.create_user(
        email='attendance-employee@test.com',
        password='pass123!',
        account_type=AccountType.WORKFORCE,
        role=UserRole.EMPLOYEE,
        organisation=organisation,
        is_active=True,
    )
    employee = Employee.objects.create(
        organisation=organisation,
        user=employee_user,
        employee_code='EMP100',
        designation='Engineer',
        status=EmployeeStatus.ACTIVE,
        date_of_joining=date(2026, 4, 1),
    )

    client = APIClient()
    client.force_authenticate(user=org_admin_user)
    session = client.session
    session['active_workspace_kind'] = 'ADMIN'
    session['active_admin_org_id'] = str(organisation.id)
    session.save()

    employee_client = APIClient()
    employee_client.force_authenticate(user=employee_user)
    employee_session = employee_client.session
    employee_session['active_workspace_kind'] = 'EMPLOYEE'
    employee_session['active_employee_org_id'] = str(organisation.id)
    employee_session.save()

    return {
        'organisation': organisation,
        'org_admin_user': org_admin_user,
        'employee_user': employee_user,
        'employee': employee,
        'client': client,
        'employee_client': employee_client,
    }


@pytest.mark.django_db
class TestAttendanceImportViews:
    def test_sample_template_downloads_return_xlsx_files(self, attendance_setup):
        client = attendance_setup['client']

        attendance_response = client.get('/api/org/attendance/imports/templates/attendance-sheet/')
        punch_response = client.get('/api/org/attendance/imports/templates/punch-sheet/')

        assert attendance_response.status_code == 200
        assert punch_response.status_code == 200
        assert attendance_response['Content-Type'].startswith('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        assert punch_response['Content-Type'].startswith('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        attendance_workbook = load_workbook(BytesIO(attendance_response.content))
        punch_workbook = load_workbook(BytesIO(punch_response.content))

        assert [attendance_workbook.active['A1'].value, attendance_workbook.active['B1'].value, attendance_workbook.active['C1'].value, attendance_workbook.active['D1'].value] == ['employee_code', 'date', 'check_in', 'check_out']
        assert [punch_workbook.active['A1'].value, punch_workbook.active['B1'].value, punch_workbook.active['C1'].value] == ['employee_code', 'date', 'punch_time']

    def test_attendance_sheet_import_posts_attendance_records(self, attendance_setup):
        client = attendance_setup['client']
        employee = attendance_setup['employee']
        workbook_bytes = build_xlsx(
            ['employee_code', 'date', 'check_in', 'check_out'],
            [['EMP100', '2026-04-02', '09:05', '18:12']],
        )

        response = client.post(
            '/api/org/attendance/imports/attendance-sheet/',
            {'file': SimpleUploadedFile('attendance.xlsx', workbook_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
        )

        assert response.status_code == 201
        assert response.data['status'] == AttendanceImportStatus.POSTED
        assert response.data['posted_rows'] == 1

        record = AttendanceRecord.objects.get(employee=employee, attendance_date=date(2026, 4, 2))
        assert record.source == 'EXCEL_IMPORT'
        assert record.check_in_at.strftime('%H:%M') == '09:05'
        assert record.check_out_at.strftime('%H:%M') == '18:12'

    def test_attendance_sheet_import_fails_without_partial_post(self, attendance_setup):
        client = attendance_setup['client']
        workbook_bytes = build_xlsx(
            ['employee_code', 'date', 'check_in', 'check_out'],
            [
                ['EMP100', '2026-04-02', '09:05', '18:12'],
                ['UNKNOWN', '2026-04-02', '09:00', '18:00'],
            ],
        )

        response = client.post(
            '/api/org/attendance/imports/attendance-sheet/',
            {'file': SimpleUploadedFile('attendance.xlsx', workbook_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
        )

        assert response.status_code == 201
        assert response.data['status'] == AttendanceImportStatus.FAILED
        assert response.data['error_rows'] == 1
        assert AttendanceRecord.objects.count() == 0

    def test_punch_sheet_import_generates_normalized_download(self, attendance_setup):
        client = attendance_setup['client']
        workbook_bytes = build_xlsx(
            ['employee_code', 'date', 'punch_time'],
            [
                ['EMP100', '2026-04-03', '09:01'],
                ['EMP100', '2026-04-03', '13:02'],
                ['EMP100', '2026-04-03', '18:04'],
            ],
        )

        upload_response = client.post(
            '/api/org/attendance/imports/punch-sheet/',
            {'file': SimpleUploadedFile('punches.xlsx', workbook_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
        )

        assert upload_response.status_code == 201
        assert upload_response.data['status'] == AttendanceImportStatus.READY_FOR_REVIEW
        assert upload_response.data['valid_rows'] == 1
        job_id = upload_response.data['id']

        download_response = client.get(f'/api/org/attendance/imports/{job_id}/normalized-file/')

        assert download_response.status_code == 200
        workbook = load_workbook(BytesIO(download_response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        assert rows[0] == ('employee_code', 'date', 'check_in', 'check_out')
        assert rows[1] == ('EMP100', '2026-04-03', '09:01', '18:04')

    def test_punch_sheet_import_marks_single_punch_days_for_review(self, attendance_setup):
        client = attendance_setup['client']
        workbook_bytes = build_xlsx(
            ['employee_code', 'date', 'punch_time'],
            [['EMP100', '2026-04-04', '09:30']],
        )

        upload_response = client.post(
            '/api/org/attendance/imports/punch-sheet/',
            {'file': SimpleUploadedFile('punches.xlsx', workbook_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
        )

        assert upload_response.status_code == 201
        assert upload_response.data['status'] == AttendanceImportStatus.READY_FOR_REVIEW

        workbook = load_workbook(BytesIO(client.get(f"/api/org/attendance/imports/{upload_response.data['id']}/normalized-file/").content))
        review_sheet = workbook['Review Notes']
        assert review_sheet['A2'].value == 'EMP100'
        assert review_sheet['D2'].value == 'Only one punch was found for this employee on this date. Add a checkout time before re-uploading.'

    def test_employee_can_punch_in_and_out_and_summary_updates(self, attendance_setup):
        client = attendance_setup['employee_client']

        punch_in_response = client.post('/api/me/attendance/punch-in/', {})
        punch_out_response = client.post('/api/me/attendance/punch-out/', {})
        summary_response = client.get('/api/me/attendance/summary/')

        assert punch_in_response.status_code == 201
        assert punch_out_response.status_code == 201
        assert summary_response.status_code == 200
        assert summary_response.data['today']['check_in_at'] is not None
        assert summary_response.data['today']['check_out_at'] is not None
        assert summary_response.data['today']['raw_punch_count'] >= 2

    def test_org_attendance_dashboard_and_days_endpoint_return_data(self, attendance_setup):
        employee_client = attendance_setup['employee_client']
        admin_client = attendance_setup['client']

        employee_client.post('/api/me/attendance/punch-in/', {})
        employee_client.post('/api/me/attendance/punch-out/', {})

        dashboard_response = admin_client.get('/api/org/attendance/dashboard/')
        days_response = admin_client.get('/api/org/attendance/days/')

        assert dashboard_response.status_code == 200
        assert days_response.status_code == 200
        assert dashboard_response.data['total_employees'] == 1
        assert len(days_response.data) == 1
        assert days_response.data[0]['employee_code'] == 'EMP100'

    def test_org_can_create_api_source_and_ingest_punches(self, attendance_setup):
        admin_client = attendance_setup['client']
        anonymous_client = APIClient()

        create_response = admin_client.post(
            '/api/org/attendance/sources/',
            {
                'name': 'Biometric gateway',
                'kind': 'API',
                'configuration': {'site': 'HQ'},
            },
            format='json',
        )

        assert create_response.status_code == 201
        assert create_response.data['raw_api_key']
        source_id = create_response.data['id']

        ingest_response = anonymous_client.post(
            f'/api/org/attendance/public-sources/{source_id}/punches/',
            {
                'punches': [
                    {'employee_code': 'EMP100', 'punch_at': '2026-04-07T09:02:00', 'action_type': 'CHECK_IN'},
                    {'employee_code': 'EMP100', 'punch_at': '2026-04-07T18:01:00', 'action_type': 'CHECK_OUT'},
                ]
            },
            format='json',
            HTTP_X_ATTENDANCE_SOURCE_KEY=create_response.data['raw_api_key'],
        )

        assert ingest_response.status_code == 202
        assert ingest_response.data['ingested_count'] == 2
        assert AttendanceSourceConfig.objects.filter(id=source_id, kind='API').exists()
        assert AttendancePunch.objects.filter(employee=attendance_setup['employee']).count() == 2
        assert AttendanceDay.objects.get(employee=attendance_setup['employee'], attendance_date=date(2026, 4, 7)).raw_punch_count == 2

    def test_org_attendance_report_returns_monthly_summary(self, attendance_setup):
        employee_client = attendance_setup['employee_client']
        admin_client = attendance_setup['client']

        employee_client.post('/api/me/attendance/punch-in/', {})
        employee_client.post('/api/me/attendance/punch-out/', {})

        response = admin_client.get('/api/org/attendance/reports/summary/', {'month': '2026-04'})

        assert response.status_code == 200
        assert response.data['month'] == '2026-04'
        assert response.data['employee_count'] == 1
        assert 'rows' in response.data

    def test_attendance_regularization_approval_updates_attendance_day(self, attendance_setup):
        organisation = attendance_setup['organisation']
        org_admin_user = attendance_setup['org_admin_user']
        employee = attendance_setup['employee']
        employee_client = attendance_setup['employee_client']

        workflow = ApprovalWorkflow.objects.create(
            organisation=organisation,
            name='Default Attendance Regularization Workflow',
            is_default=True,
            default_request_kind=ApprovalRequestKind.ATTENDANCE_REGULARIZATION,
            is_active=True,
        )
        ApprovalWorkflowRule.objects.create(
            workflow=workflow,
            name='Default regularization rule',
            request_kind=ApprovalRequestKind.ATTENDANCE_REGULARIZATION,
            priority=100,
            is_active=True,
        )
        stage = ApprovalStage.objects.create(
            workflow=workflow,
            name='Org admin review',
            sequence=1,
            mode='ALL',
        )
        ApprovalStageApprover.objects.create(
            stage=stage,
            approver_type=ApprovalApproverType.PRIMARY_ORG_ADMIN,
        )

        response = employee_client.post(
            '/api/me/attendance/regularizations/',
            {
                'attendance_date': '2026-04-06',
                'requested_check_in': '09:10',
                'requested_check_out': '18:05',
                'reason': 'Biometric sync missed both punches.',
            },
            format='json',
        )

        assert response.status_code == 201
        regularization = AttendanceRegularizationRequest.objects.get(id=response.data['id'])
        action = regularization.approval_run.actions.get()
        approve_action(action, org_admin_user)

        regularization.refresh_from_db()
        day = AttendanceDay.objects.get(employee=employee, attendance_date=date(2026, 4, 6))

        assert regularization.status == AttendanceRegularizationStatus.APPROVED
        assert day.check_in_at is not None
        assert day.check_out_at is not None
        assert day.status in {'PRESENT', 'HALF_DAY'}

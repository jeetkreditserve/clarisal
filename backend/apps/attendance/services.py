from __future__ import annotations

import ipaddress
import math
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.common.security import decrypt_value, encrypt_value, generate_secure_token, hash_token, mask_value
from apps.approvals.models import ApprovalRequestKind, ApprovalRun
from apps.approvals.services import cancel_approval_run, create_approval_run
from apps.audit.services import log_audit_event
from apps.employees.models import Employee, EmployeeStatus
from apps.timeoff.models import (
    DaySession,
    Holiday,
    HolidayCalendar,
    HolidayCalendarStatus,
    LeaveRequest,
    LeaveRequestStatus,
    OnDutyDurationType,
    OnDutyRequest,
    OnDutyRequestStatus,
)

from .models import (
    AttendanceDay,
    AttendanceDayStatus,
    AttendanceImportJob,
    AttendanceImportMode,
    AttendanceImportRow,
    AttendanceImportRowStatus,
    AttendanceImportStatus,
    AttendancePolicy,
    AttendancePunch,
    AttendancePunchActionType,
    AttendancePunchSource,
    AttendanceRecord,
    AttendanceRecordSource,
    AttendanceRegularizationRequest,
    AttendanceRegularizationStatus,
    AttendanceSourceConfig,
    AttendanceSourceConfigKind,
    Shift,
    ShiftAssignment,
)

ATTENDANCE_HEADERS = ['employee_code', 'date', 'check_in', 'check_out']
PUNCH_HEADERS = ['employee_code', 'date', 'punch_time']
XLSX_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
ZERO = Decimal('0.00')
HALF = Decimal('0.50')
FULL = Decimal('1.00')
DEFAULT_WEEK_OFF_DAYS = [6]


def _decimal(value):
    return Decimal(str(value)).quantize(Decimal('0.01'))


def _normalize_header(value):
    return str(value or '').strip().lower().replace(' ', '_')


def _strip_string(value):
    return str(value or '').strip()


def _parse_date(value):
    if value is None or value == '':
        raise ValueError('Date is required.')
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _strip_string(value)
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError('Date must be a valid Excel date or YYYY-MM-DD style value.')


def _parse_time_value(value, *, label):
    if value is None or value == '':
        raise ValueError(f'{label} is required.')
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    text = _strip_string(value)
    for fmt in (
        '%H:%M',
        '%H:%M:%S',
        '%I:%M %p',
        '%I:%M:%S %p',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d %H:%M:%S',
        '%d-%m-%Y %H:%M',
        '%d-%m-%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M',
        '%Y-%m-%dT%H:%M:%S',
    ):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError(f'{label} must be a valid Excel time or time-like value.')


def _policy_timezone(policy):
    try:
        return ZoneInfo(policy.timezone_name)
    except Exception:  # noqa: BLE001
        return timezone.get_default_timezone()


def _policy_local_date(policy, dt=None):
    current_dt = dt or timezone.now()
    return timezone.localtime(current_dt, _policy_timezone(policy)).date()


def _make_aware_datetime(attendance_date, time_value, policy=None):
    naive = datetime.combine(attendance_date, time_value)
    return timezone.make_aware(naive, _policy_timezone(policy) if policy else timezone.get_default_timezone())


def _parse_datetime_value(value, *, label, policy=None):
    if value in (None, ''):
        raise ValueError(f'{label} is required.')
    if isinstance(value, datetime):
        return value if timezone.is_aware(value) else timezone.make_aware(value, _policy_timezone(policy) if policy else timezone.get_default_timezone())
    text = _strip_string(value)
    for fmt in (
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%dT%H:%M',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%d-%m-%Y %H:%M:%S',
        '%d-%m-%Y %H:%M',
    ):
        try:
            parsed = datetime.strptime(text, fmt)
            return timezone.make_aware(parsed, _policy_timezone(policy) if policy else timezone.get_default_timezone())
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        return parsed if timezone.is_aware(parsed) else timezone.make_aware(parsed, _policy_timezone(policy) if policy else timezone.get_default_timezone())
    except ValueError as exc:
        raise ValueError(f'{label} must be a valid ISO datetime.') from exc


def _load_sheet_rows(uploaded_file, expected_headers):
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError('The uploaded workbook is empty.')

    headers = [_normalize_header(value) for value in rows[0]]
    if headers[: len(expected_headers)] != expected_headers:
        raise ValueError(f'The workbook must contain these columns in order: {", ".join(expected_headers)}.')

    parsed_rows = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, '') for cell in row[: len(expected_headers)]):
            continue
        parsed_rows.append(
            {
                expected_headers[index]: row[index] if index < len(row) else None
                for index in range(len(expected_headers))
            }
            | {'row_number': row_number}
        )
    return parsed_rows


def _resolve_employee(organisation, employee_code):
    normalized_code = _strip_string(employee_code).upper()
    if not normalized_code:
        raise ValueError('Employee code is required.')
    employee = Employee.objects.filter(
        organisation=organisation,
        employee_code=normalized_code,
        status=EmployeeStatus.ACTIVE,
    ).select_related('user', 'office_location').first()
    if employee is None:
        raise ValueError('Employee code was not found in this organisation or is not active.')
    return normalized_code, employee


def _build_workbook_bytes(title, headers, rows, *, instructions=None, review_notes=None, validation_errors=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    if instructions:
        instructions_sheet = workbook.create_sheet('Instructions')
        instructions_sheet.append(['Instruction'])
        for item in instructions:
            instructions_sheet.append([item])

    if review_notes:
        notes_sheet = workbook.create_sheet('Review Notes')
        notes_sheet.append(['employee_code', 'date', 'punch_count', 'review_note'])
        for row in review_notes:
            notes_sheet.append(row)

    if validation_errors:
        errors_sheet = workbook.create_sheet('Validation Errors')
        errors_sheet.append(['row_number', 'employee_code', 'message'])
        for row in validation_errors:
            errors_sheet.append(row)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_attendance_sheet_sample():
    return _build_workbook_bytes(
        'Attendance Import',
        ATTENDANCE_HEADERS,
        [
            ['EMP100', '2026-04-01', '09:02', '18:11'],
            ['EMP101', '2026-04-01', '09:15', '18:07'],
        ],
        instructions=[
            'Use one row per employee per date.',
            'Required columns: employee_code, date, check_in, check_out.',
            'The file posts attendance directly after validation.',
            'Dates should be in YYYY-MM-DD or Excel date format.',
            'Times should be in HH:MM, HH:MM:SS, or Excel time format.',
        ],
    )


def build_punch_sheet_sample():
    return _build_workbook_bytes(
        'Punch Import',
        PUNCH_HEADERS,
        [
            ['EMP100', '2026-04-01', '09:02'],
            ['EMP100', '2026-04-01', '18:11'],
            ['EMP101', '2026-04-01', '09:15'],
        ],
        instructions=[
            'Use one row per punch.',
            'Required columns: employee_code, date, punch_time.',
            'The system converts each employee-day into one attendance row using first punch as check-in and last punch as check-out.',
            'Single-punch days are returned with a blank checkout for review.',
        ],
    )


def build_normalized_attendance_workbook(job):
    policy = get_default_attendance_policy(job.organisation)
    policy_timezone = _policy_timezone(policy)
    normalized_rows = []
    review_notes = []
    validation_errors = []
    for row in job.rows.order_by('row_number', 'created_at'):
        if row.status in {AttendanceImportRowStatus.VALID, AttendanceImportRowStatus.INCOMPLETE}:
            normalized_rows.append(
                [
                    row.employee_code,
                    row.attendance_date.isoformat() if row.attendance_date else '',
                    timezone.localtime(row.check_in_at, policy_timezone).strftime('%H:%M') if row.check_in_at else '',
                    timezone.localtime(row.check_out_at, policy_timezone).strftime('%H:%M') if row.check_out_at else '',
                ]
            )
            review_notes.append(
                [
                    row.employee_code,
                    row.attendance_date.isoformat() if row.attendance_date else '',
                    len(row.raw_punch_times),
                    row.error_message or 'Review generated attendance before re-uploading.',
                ]
            )
        elif row.status == AttendanceImportRowStatus.ERROR:
            validation_errors.append([row.row_number, row.employee_code, row.error_message])

    return _build_workbook_bytes(
        'Attendance Import',
        ATTENDANCE_HEADERS,
        normalized_rows,
        instructions=[
            'This file was generated from a raw punch upload.',
            'Review every row before re-uploading it through the attendance-sheet import.',
            'Blank checkout values must be corrected before posting attendance.',
        ],
        review_notes=review_notes,
        validation_errors=validation_errors,
    )


def _ensure_xlsx(uploaded_file):
    filename = getattr(uploaded_file, 'name', '')
    if not filename.lower().endswith('.xlsx'):
        raise ValueError('Only Excel .xlsx files are supported for attendance import.')


def get_default_attendance_policy(organisation):
    policy = AttendancePolicy.objects.filter(
        organisation=organisation,
        is_default=True,
        is_active=True,
    ).first()
    if policy is None:
        policy = AttendancePolicy.objects.create(
            organisation=organisation,
            name='Default Attendance Policy',
            week_off_days=DEFAULT_WEEK_OFF_DAYS,
            is_default=True,
            is_active=True,
        )
    return policy


def upsert_attendance_policy(organisation, *, actor=None, policy=None, **fields):
    if policy is None:
        policy = AttendancePolicy.objects.create(
            organisation=organisation,
            week_off_days=fields.pop('week_off_days', DEFAULT_WEEK_OFF_DAYS),
            **fields,
        )
    else:
        for attr, value in fields.items():
            setattr(policy, attr, value)
        if not policy.week_off_days:
            policy.week_off_days = DEFAULT_WEEK_OFF_DAYS
        policy.save()

    if policy.is_default:
        AttendancePolicy.objects.filter(organisation=organisation).exclude(id=policy.id).update(is_default=False)
    log_audit_event(actor, 'attendance.policy.upserted', organisation=organisation, target=policy)
    return policy


def create_source_config(organisation, *, name, kind, actor=None, configuration=None, is_active=True):
    config = dict(configuration or {})
    raw_api_key = None
    if kind == AttendanceSourceConfigKind.API:
        raw_api_key = config.pop('raw_api_key', None) or generate_secure_token(24)
        config['api_key_hash'] = hash_token(raw_api_key)
        config['api_key_encrypted'] = encrypt_value(raw_api_key)
    source = AttendanceSourceConfig.objects.create(
        organisation=organisation,
        name=name,
        kind=kind,
        configuration=config,
        is_active=is_active,
        created_by=actor,
        modified_by=actor,
    )
    log_audit_event(actor, 'attendance.source.created', organisation=organisation, target=source, payload={'kind': kind})
    return source, raw_api_key


def update_source_config(source, *, actor=None, name=None, is_active=None, configuration=None, rotate_api_key=False):
    changed = False
    raw_api_key = None

    if name is not None and source.name != name:
        source.name = name
        changed = True
    if is_active is not None and source.is_active != is_active:
        source.is_active = is_active
        changed = True

    config = dict(source.configuration or {})
    if configuration:
        config.update(configuration)
        changed = True

    if source.kind == AttendanceSourceConfigKind.API and rotate_api_key:
        raw_api_key = generate_secure_token(24)
        config['api_key_hash'] = hash_token(raw_api_key)
        config['api_key_encrypted'] = encrypt_value(raw_api_key)
        changed = True

    if changed:
        source.configuration = config
        source.modified_by = actor
        source.save(update_fields=['name', 'is_active', 'configuration', 'modified_by', 'modified_at'])
        log_audit_event(actor, 'attendance.source.updated', organisation=source.organisation, target=source, payload={'rotated_api_key': rotate_api_key})
    return source, raw_api_key


def get_source_api_key_preview(source):
    encrypted = (source.configuration or {}).get('api_key_encrypted', '')
    raw = decrypt_value(encrypted)
    return mask_value(raw)


def create_shift(organisation, *, actor=None, **fields):
    shift = Shift.objects.create(organisation=organisation, **fields)
    log_audit_event(actor, 'attendance.shift.created', organisation=organisation, target=shift)
    return shift


def update_shift(shift, *, actor=None, **fields):
    for attr, value in fields.items():
        setattr(shift, attr, value)
    shift.save()
    log_audit_event(actor, 'attendance.shift.updated', organisation=shift.organisation, target=shift)
    return shift


def assign_shift(employee, shift, *, start_date, end_date=None, actor=None):
    assignment = ShiftAssignment.objects.create(
        organisation=employee.organisation,
        employee=employee,
        shift=shift,
        start_date=start_date,
        end_date=end_date,
    )
    log_audit_event(actor, 'attendance.shift.assigned', organisation=employee.organisation, target=assignment)
    return assignment


def _get_effective_shift(employee, attendance_date):
    return employee.shift_assignments.filter(
        organisation=employee.organisation,
        is_active=True,
        start_date__lte=attendance_date,
    ).filter(
        Q(end_date__isnull=True) | Q(end_date__gte=attendance_date)
    ).select_related('shift').order_by('-start_date', '-created_at').first()


def _get_shift_windows(policy, shift, attendance_date):
    start_time_value = shift.start_time if shift else policy.default_start_time
    end_time_value = shift.end_time if shift else policy.default_end_time
    tz = _policy_timezone(policy)
    shift_start = timezone.make_aware(datetime.combine(attendance_date, start_time_value), tz)
    if shift and shift.is_overnight:
        shift_end = timezone.make_aware(datetime.combine(attendance_date + timedelta(days=1), end_time_value), tz)
    else:
        shift_end = timezone.make_aware(datetime.combine(attendance_date, end_time_value), tz)
        if shift_end < shift_start:
            shift_end = timezone.make_aware(datetime.combine(attendance_date + timedelta(days=1), end_time_value), tz)
    return shift_start, shift_end


def _get_threshold(policy, shift, attr):
    shift_value = getattr(shift, attr) if shift is not None else None
    if shift_value is not None:
        return shift_value
    return getattr(policy, attr)


def _weekday_week_off(policy, attendance_date):
    week_off_days = policy.week_off_days or DEFAULT_WEEK_OFF_DAYS
    return attendance_date.weekday() in week_off_days


def _get_holiday_for_date(employee, attendance_date):
    published_calendars = HolidayCalendar.objects.filter(
        organisation=employee.organisation,
        status=HolidayCalendarStatus.PUBLISHED,
        year=attendance_date.year,
    ).prefetch_related('location_assignments')
    employee_location_id = employee.office_location_id
    for calendar_obj in published_calendars:
        assigned_locations = set(calendar_obj.location_assignments.values_list('office_location_id', flat=True))
        if assigned_locations and employee_location_id not in assigned_locations:
            continue
        holiday = calendar_obj.holidays.filter(holiday_date=attendance_date).first()
        if holiday:
            return holiday
    return None


def _request_day_fraction(start_date, end_date, attendance_date, start_session, end_session):
    if attendance_date < start_date or attendance_date > end_date:
        return ZERO
    if start_date == end_date == attendance_date:
        if start_session != DaySession.FULL_DAY or end_session != DaySession.FULL_DAY:
            return HALF
        return FULL
    if attendance_date == start_date and start_session != DaySession.FULL_DAY:
        return HALF
    if attendance_date == end_date and end_session != DaySession.FULL_DAY:
        return HALF
    return FULL


def _get_leave_fraction(employee, attendance_date):
    total = ZERO
    for request in LeaveRequest.objects.filter(
        employee=employee,
        status=LeaveRequestStatus.APPROVED,
        start_date__lte=attendance_date,
        end_date__gte=attendance_date,
    ):
        total += _request_day_fraction(
            request.start_date,
            request.end_date,
            attendance_date,
            request.start_session,
            request.end_session,
        )
    return min(total, FULL)


def _get_on_duty_fraction(employee, attendance_date, half_day_minutes):
    total = ZERO
    for request in OnDutyRequest.objects.filter(
        employee=employee,
        status=OnDutyRequestStatus.APPROVED,
        start_date__lte=attendance_date,
        end_date__gte=attendance_date,
    ):
        if request.duration_type == OnDutyDurationType.TIME_RANGE and request.start_date == request.end_date == attendance_date:
            if request.start_time and request.end_time:
                start_dt = datetime.combine(attendance_date, request.start_time)
                end_dt = datetime.combine(attendance_date, request.end_time)
                duration_minutes = max(0, int((end_dt - start_dt).total_seconds() // 60))
                total += FULL if duration_minutes >= half_day_minutes else HALF
            continue
        if request.duration_type == OnDutyDurationType.FULL_DAY:
            total += FULL
        elif request.duration_type in {OnDutyDurationType.FIRST_HALF, OnDutyDurationType.SECOND_HALF}:
            total += HALF
        else:
            total += _request_day_fraction(
                request.start_date,
                request.end_date,
                attendance_date,
                DaySession.FULL_DAY,
                DaySession.FULL_DAY,
            )
    return min(total, FULL)


def _get_day_bounds(policy, shift, attendance_date):
    shift_start, shift_end = _get_shift_windows(policy, shift, attendance_date)
    if shift and shift.is_overnight:
        return shift_start - timedelta(hours=6), shift_end + timedelta(hours=6), shift_start, shift_end
    day_start = timezone.make_aware(datetime.combine(attendance_date, time.min), _policy_timezone(policy))
    day_end = timezone.make_aware(datetime.combine(attendance_date, time.max), _policy_timezone(policy))
    return day_start, day_end, shift_start, shift_end


def _pick_interval_from_punches(punches):
    if not punches:
        return None, None
    check_in_candidates = [item.punch_at for item in punches if item.action_type == AttendancePunchActionType.CHECK_IN]
    check_out_candidates = [item.punch_at for item in punches if item.action_type == AttendancePunchActionType.CHECK_OUT]
    if check_in_candidates:
        check_in_at = min(check_in_candidates)
    else:
        check_in_at = punches[0].punch_at

    valid_check_outs = [item for item in check_out_candidates if item >= check_in_at]
    if valid_check_outs:
        check_out_at = max(valid_check_outs)
    elif len(punches) > 1:
        check_out_at = punches[-1].punch_at if punches[-1].punch_at > check_in_at else None
    else:
        check_out_at = None
    return check_in_at, check_out_at


def _get_attendance_override(employee, attendance_date):
    return AttendanceRecord.objects.filter(
        organisation=employee.organisation,
        employee=employee,
        attendance_date=attendance_date,
    ).first()


def _summarize_day(employee, attendance_date):
    policy = get_default_attendance_policy(employee.organisation)
    assignment = _get_effective_shift(employee, attendance_date)
    shift = assignment.shift if assignment else None
    day_start, day_end, shift_start, _shift_end = _get_day_bounds(policy, shift, attendance_date)
    half_day_min_minutes = _get_threshold(policy, shift, 'half_day_min_minutes')
    full_day_min_minutes = _get_threshold(policy, shift, 'full_day_min_minutes')
    overtime_after_minutes = _get_threshold(policy, shift, 'overtime_after_minutes')
    grace_minutes = _get_threshold(policy, shift, 'grace_minutes')

    holiday = _get_holiday_for_date(employee, attendance_date)
    leave_fraction = _get_leave_fraction(employee, attendance_date)
    on_duty_fraction = _get_on_duty_fraction(employee, attendance_date, half_day_min_minutes)
    override_record = _get_attendance_override(employee, attendance_date)
    punches = list(
        AttendancePunch.objects.filter(
            organisation=employee.organisation,
            employee=employee,
            punch_at__gte=day_start,
            punch_at__lte=day_end,
        ).order_by('punch_at')
    )

    check_in_at = None
    check_out_at = None
    source = ''
    note = ''
    if override_record is not None:
        check_in_at = override_record.check_in_at
        check_out_at = override_record.check_out_at
        source = (
            AttendancePunchSource.REGULARIZATION
            if override_record.source == AttendanceRecordSource.REGULARIZATION
            else AttendancePunchSource.MANUAL
            if override_record.source == AttendanceRecordSource.MANUAL_OVERRIDE
            else AttendancePunchSource.IMPORT
        )
    elif punches:
        check_in_at, check_out_at = _pick_interval_from_punches(punches)
        source = punches[-1].source

    worked_minutes = 0
    late_minutes = 0
    overtime_minutes = 0
    is_late = False
    needs_regularization = False
    paid_fraction = ZERO
    status = AttendanceDayStatus.ABSENT

    if check_in_at and check_out_at:
        worked_minutes = max(0, int((check_out_at - check_in_at).total_seconds() // 60))
        if worked_minutes >= full_day_min_minutes:
            status = AttendanceDayStatus.PRESENT
            paid_fraction = FULL
        elif worked_minutes >= half_day_min_minutes:
            status = AttendanceDayStatus.HALF_DAY
            paid_fraction = HALF
        else:
            status = AttendanceDayStatus.ABSENT
            paid_fraction = ZERO

        late_threshold = shift_start + timedelta(minutes=grace_minutes)
        if check_in_at > late_threshold:
            late_minutes = max(0, int((check_in_at - late_threshold).total_seconds() // 60))
            is_late = True
        if worked_minutes > overtime_after_minutes:
            overtime_minutes = worked_minutes - overtime_after_minutes
    elif check_in_at and not check_out_at:
        status = AttendanceDayStatus.INCOMPLETE
        needs_regularization = True
        note = 'Only one punch is available for this day.'
    else:
        if holiday:
            status = AttendanceDayStatus.HOLIDAY
            paid_fraction = FULL
        elif _weekday_week_off(policy, attendance_date):
            status = AttendanceDayStatus.WEEK_OFF
            paid_fraction = FULL

    if leave_fraction >= FULL:
        status = AttendanceDayStatus.ON_LEAVE
        paid_fraction = FULL
        needs_regularization = False
    elif on_duty_fraction >= FULL and status in {AttendanceDayStatus.ABSENT, AttendanceDayStatus.INCOMPLETE}:
        status = AttendanceDayStatus.ON_DUTY
        paid_fraction = FULL
        needs_regularization = False
    elif leave_fraction == HALF:
        paid_fraction = min(FULL, paid_fraction + HALF)
        status = AttendanceDayStatus.PRESENT if paid_fraction >= FULL else AttendanceDayStatus.HALF_DAY
        needs_regularization = False
    elif on_duty_fraction == HALF and status in {AttendanceDayStatus.ABSENT, AttendanceDayStatus.HALF_DAY, AttendanceDayStatus.INCOMPLETE}:
        paid_fraction = min(FULL, paid_fraction + HALF)
        status = AttendanceDayStatus.PRESENT if paid_fraction >= FULL else AttendanceDayStatus.HALF_DAY
        needs_regularization = False

    return {
        'policy': policy,
        'shift': shift,
        'status': status,
        'source': source,
        'check_in_at': check_in_at,
        'check_out_at': check_out_at,
        'worked_minutes': worked_minutes,
        'overtime_minutes': overtime_minutes,
        'late_minutes': late_minutes,
        'is_late': is_late,
        'needs_regularization': needs_regularization,
        'paid_fraction': _decimal(paid_fraction),
        'leave_fraction': _decimal(leave_fraction),
        'on_duty_fraction': _decimal(on_duty_fraction),
        'is_holiday': holiday is not None,
        'is_week_off': _weekday_week_off(policy, attendance_date),
        'raw_punch_count': len(punches),
        'note': note,
        'metadata': {
            'holiday_name': holiday.name if holiday else '',
            'override_source': override_record.source if override_record else '',
        },
    }


def recalculate_attendance_day(employee, attendance_date, *, actor=None):
    summary = _summarize_day(employee, attendance_date)
    attendance_day, _ = AttendanceDay.objects.update_or_create(
        organisation=employee.organisation,
        employee=employee,
        attendance_date=attendance_date,
        defaults=summary,
    )
    if actor is not None:
        log_audit_event(
            actor,
            'attendance.day.recalculated',
            organisation=employee.organisation,
            target=attendance_day,
            payload={'attendance_date': attendance_date.isoformat()},
        )
    return attendance_day


def get_employee_attendance_summary(employee):
    policy = get_default_attendance_policy(employee.organisation)
    today = _policy_local_date(policy)
    attendance_day = recalculate_attendance_day(employee, today)
    shift_assignment = _get_effective_shift(employee, today)
    pending_regularizations = employee.attendance_regularization_requests.filter(
        status=AttendanceRegularizationStatus.PENDING
    ).order_by('-created_at')[:5]
    return {
        'today': attendance_day,
        'policy': policy,
        'shift': shift_assignment.shift if shift_assignment else None,
        'pending_regularizations': list(pending_regularizations),
    }


def _month_window(month, *, policy=None):
    if month:
        year, month_number = [int(part) for part in month.split('-')]
        start_date = date(year, month_number, 1)
    else:
        today = _policy_local_date(policy) if policy is not None else timezone.localdate()
        start_date = date(today.year, today.month, 1)
    next_month = date(start_date.year + (1 if start_date.month == 12 else 0), 1 if start_date.month == 12 else start_date.month + 1, 1)
    end_date = next_month - timedelta(days=1)
    return start_date, end_date


def get_employee_attendance_history(employee, *, month=None):
    start_date, end_date = _month_window(month, policy=get_default_attendance_policy(employee.organisation))
    days = []
    current = start_date
    while current <= end_date:
        days.append(recalculate_attendance_day(employee, current))
        current += timedelta(days=1)
    return days


def get_employee_attendance_calendar(employee, *, month=None):
    days = get_employee_attendance_history(employee, month=month)
    return {
        'month': month or _month_window(None, policy=get_default_attendance_policy(employee.organisation))[0].strftime('%Y-%m'),
        'days': [
            {
                'date': day.attendance_date.isoformat(),
                'status': day.status,
                'is_late': day.is_late,
                'needs_regularization': day.needs_regularization,
                'worked_minutes': day.worked_minutes,
                'overtime_minutes': day.overtime_minutes,
            }
            for day in days
        ],
    }


def _extract_remote_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return _strip_string(request.META.get('REMOTE_ADDR'))


def _validate_ip(policy, remote_ip):
    if not policy.restrict_by_ip:
        return
    if not remote_ip:
        raise ValueError('This attendance policy allows punches only from approved IP ranges.')
    for value in policy.allowed_ip_ranges:
        try:
            network = ipaddress.ip_network(value, strict=False)
            if ipaddress.ip_address(remote_ip) in network:
                return
        except ValueError:
            continue
    raise ValueError('Your current network is not allowed by the attendance policy.')


def _haversine_distance_meters(lat1, lon1, lat2, lon2):
    earth_radius = 6371000
    phi1 = math.radians(float(lat1))
    phi2 = math.radians(float(lat2))
    delta_phi = math.radians(float(lat2) - float(lat1))
    delta_lambda = math.radians(float(lon2) - float(lon1))
    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return earth_radius * c


def _validate_geo(policy, latitude, longitude):
    if not policy.restrict_by_geo:
        return
    if latitude is None or longitude is None:
        raise ValueError('This attendance policy requires a location-enabled punch.')
    for site in policy.allowed_geo_sites:
        try:
            distance = _haversine_distance_meters(latitude, longitude, site['latitude'], site['longitude'])
            if distance <= float(site.get('radius_meters', 250)):
                return
        except Exception:  # noqa: BLE001
            continue
    raise ValueError('You are outside the approved attendance location.')


def record_employee_punch(employee, *, action_type, actor=None, remote_ip='', latitude=None, longitude=None, source=AttendancePunchSource.WEB):
    policy = get_default_attendance_policy(employee.organisation)
    if source == AttendancePunchSource.WEB and not policy.allow_web_punch:
        raise ValueError('Web punch is disabled for this attendance policy.')
    _validate_ip(policy, remote_ip)
    _validate_geo(policy, latitude, longitude)

    today = _policy_local_date(policy)
    attendance_day = recalculate_attendance_day(employee, today)
    if action_type == AttendancePunchActionType.CHECK_IN and attendance_day.check_in_at and not attendance_day.check_out_at:
        raise ValueError('You are already checked in for today.')
    if action_type == AttendancePunchActionType.CHECK_OUT and not attendance_day.check_in_at:
        raise ValueError('Check in before punching out.')
    punch = AttendancePunch.objects.create(
        organisation=employee.organisation,
        employee=employee,
        action_type=action_type,
        source=source,
        punch_at=timezone.now(),
        remote_ip=remote_ip,
        latitude=latitude,
        longitude=longitude,
    )
    attendance_day = recalculate_attendance_day(employee, today, actor=actor)
    log_audit_event(actor or employee.user, 'attendance.punch.recorded', organisation=employee.organisation, target=punch, payload={'action_type': action_type})
    return punch, attendance_day


def ingest_source_punches(source_config, *, punches, actor=None):
    if not source_config.is_active:
        raise ValueError('This attendance source is inactive.')
    organisation = source_config.organisation
    policy = get_default_attendance_policy(organisation)
    ingested = []
    duplicates = 0
    errors = []

    for index, payload in enumerate(punches, start=1):
        employee_code = _strip_string(payload.get('employee_code')).upper()
        try:
            _normalized_code, employee = _resolve_employee(organisation, employee_code)
            punch_at = _parse_datetime_value(payload.get('punch_at'), label='Punch timestamp', policy=policy)
            action_type = payload.get('action_type') or AttendancePunchActionType.RAW
            if action_type not in AttendancePunchActionType.values:
                raise ValueError('Action type must be CHECK_IN, CHECK_OUT, or RAW.')
            if AttendancePunch.objects.filter(
                organisation=organisation,
                employee=employee,
                source_config=source_config,
                punch_at=punch_at,
                action_type=action_type,
            ).exists():
                duplicates += 1
                continue
            punch = AttendancePunch.objects.create(
                organisation=organisation,
                employee=employee,
                source_config=source_config,
                action_type=action_type,
                source=AttendancePunchSource.API if source_config.kind == AttendanceSourceConfigKind.API else AttendancePunchSource.MANUAL,
                punch_at=punch_at,
                remote_ip=_strip_string(payload.get('remote_ip')),
                latitude=payload.get('latitude'),
                longitude=payload.get('longitude'),
                metadata={
                    'external_reference': _strip_string(payload.get('external_reference')),
                    'batch_id': _strip_string(payload.get('batch_id')),
                },
            )
            recalculate_attendance_day(employee, timezone.localtime(punch_at).date(), actor=actor)
            ingested.append(punch)
        except ValueError as exc:
            errors.append({'row_number': index, 'employee_code': employee_code, 'message': str(exc)})

    log_audit_event(
        actor,
        'attendance.source.ingested',
        organisation=organisation,
        target=source_config,
        payload={'received': len(punches), 'ingested': len(ingested), 'duplicates': duplicates, 'errors': len(errors)},
    )
    return {
        'received_count': len(punches),
        'ingested_count': len(ingested),
        'duplicate_count': duplicates,
        'error_count': len(errors),
        'errors': errors[:25],
    }


def upsert_attendance_override(employee, attendance_date, *, check_in_at=None, check_out_at=None, source=AttendanceRecordSource.MANUAL_OVERRIDE, actor=None, note=''):
    if check_in_at is None and check_out_at is not None:
        check_in_at = check_out_at
    if check_in_at is None:
        raise ValueError('A check-in time is required to save an attendance override.')
    record, _ = AttendanceRecord.objects.update_or_create(
        organisation=employee.organisation,
        employee=employee,
        attendance_date=attendance_date,
        defaults={
            'check_in_at': check_in_at,
            'check_out_at': check_out_at,
            'source': source,
        },
    )
    attendance_day = recalculate_attendance_day(employee, attendance_date, actor=actor)
    if note:
        attendance_day.note = note
        attendance_day.save(update_fields=['note', 'modified_at'])
    log_audit_event(actor, 'attendance.override.upserted', organisation=employee.organisation, target=record, payload={'source': source})
    return attendance_day


def create_regularization_request(employee, *, attendance_date, requested_check_in_at=None, requested_check_out_at=None, reason, actor=None):
    if requested_check_in_at is None and requested_check_out_at is None:
        raise ValueError('Provide a check-in time, a check-out time, or both.')
    attendance_day = recalculate_attendance_day(employee, attendance_date)
    if employee.attendance_regularization_requests.filter(
        attendance_date=attendance_date,
        status=AttendanceRegularizationStatus.PENDING,
    ).exists():
        raise ValueError('A pending regularization request already exists for this date.')

    with transaction.atomic():
        regularization = AttendanceRegularizationRequest.objects.create(
            organisation=employee.organisation,
            employee=employee,
            attendance_day=attendance_day,
            attendance_date=attendance_date,
            requested_check_in_at=requested_check_in_at,
            requested_check_out_at=requested_check_out_at,
            reason=reason,
        )
        approval_run = create_approval_run(
            regularization,
            ApprovalRequestKind.ATTENDANCE_REGULARIZATION,
            employee,
            actor=actor or employee.user,
            subject_label=f'Attendance regularization for {attendance_date.isoformat()}',
        )
        regularization.approval_run = approval_run
        regularization.save(update_fields=['approval_run', 'modified_at'])

    log_audit_event(actor or employee.user, 'attendance.regularization.created', organisation=employee.organisation, target=regularization)
    return regularization


def withdraw_regularization_request(regularization, *, actor=None):
    if regularization.status != AttendanceRegularizationStatus.PENDING:
        raise ValueError('Only pending regularization requests can be withdrawn.')
    regularization.status = AttendanceRegularizationStatus.WITHDRAWN
    regularization.save(update_fields=['status', 'modified_at'])
    if regularization.approval_run_id:
        cancel_approval_run(regularization.approval_run, actor=actor)
    log_audit_event(actor, 'attendance.regularization.withdrawn', organisation=regularization.organisation, target=regularization)
    return regularization


def apply_regularization_status_change(regularization, new_status, rejection_reason=''):
    update_fields = ['status', 'modified_at']
    regularization.status = new_status
    if rejection_reason:
        regularization.rejection_reason = rejection_reason
        update_fields.append('rejection_reason')
    regularization.save(update_fields=update_fields)

    if new_status == AttendanceRegularizationStatus.APPROVED:
        existing_day = recalculate_attendance_day(regularization.employee, regularization.attendance_date)
        check_in_at = regularization.requested_check_in_at or existing_day.check_in_at or regularization.requested_check_out_at
        check_out_at = regularization.requested_check_out_at or existing_day.check_out_at
        upsert_attendance_override(
            regularization.employee,
            regularization.attendance_date,
            check_in_at=check_in_at,
            check_out_at=check_out_at,
            source=AttendanceRecordSource.REGULARIZATION,
            note='Approved attendance regularization applied.',
        )
    else:
        recalculate_attendance_day(regularization.employee, regularization.attendance_date)
    return regularization


def get_org_attendance_dashboard(organisation, *, target_date=None):
    day = target_date or _policy_local_date(get_default_attendance_policy(organisation))
    employees = Employee.objects.filter(organisation=organisation, status=EmployeeStatus.ACTIVE).select_related('user', 'office_location')
    summaries = [recalculate_attendance_day(employee, day) for employee in employees]
    return {
        'date': day.isoformat(),
        'total_employees': len(summaries),
        'present_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.PRESENT),
        'half_day_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.HALF_DAY),
        'absent_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.ABSENT),
        'incomplete_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.INCOMPLETE),
        'holiday_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.HOLIDAY),
        'week_off_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.WEEK_OFF),
        'on_leave_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.ON_LEAVE),
        'on_duty_count': sum(1 for item in summaries if item.status == AttendanceDayStatus.ON_DUTY),
        'pending_regularizations': AttendanceRegularizationRequest.objects.filter(
            organisation=organisation,
            status=AttendanceRegularizationStatus.PENDING,
        ).count(),
        'days': summaries[:20],
    }


def get_org_attendance_report(organisation, *, month=None):
    if month:
        try:
            year, month_value = [int(part) for part in month.split('-', maxsplit=1)]
            report_month = date(year, month_value, 1)
        except (TypeError, ValueError) as exc:
            raise ValueError('month must use YYYY-MM format.') from exc
    else:
        today = _policy_local_date(get_default_attendance_policy(organisation))
        report_month = date(today.year, today.month, 1)

    next_month = date(report_month.year + (1 if report_month.month == 12 else 0), 1 if report_month.month == 12 else report_month.month + 1, 1)
    period_end = next_month - timedelta(days=1)
    employees = list(Employee.objects.filter(organisation=organisation, status=EmployeeStatus.ACTIVE).select_related('user'))
    summary = {
        'month': report_month.strftime('%Y-%m'),
        'employee_count': len(employees),
        'present_days': 0,
        'half_days': 0,
        'absent_days': 0,
        'incomplete_days': 0,
        'late_marks': 0,
        'overtime_minutes': 0,
        'rows': [],
    }

    current = report_month
    while current <= period_end:
        for employee in employees:
            day = recalculate_attendance_day(employee, current)
            if day.status == AttendanceDayStatus.PRESENT:
                summary['present_days'] += 1
            elif day.status == AttendanceDayStatus.HALF_DAY:
                summary['half_days'] += 1
            elif day.status == AttendanceDayStatus.ABSENT:
                summary['absent_days'] += 1
            elif day.status == AttendanceDayStatus.INCOMPLETE:
                summary['incomplete_days'] += 1
            if day.is_late:
                summary['late_marks'] += 1
            summary['overtime_minutes'] += day.overtime_minutes
            if day.status in {AttendanceDayStatus.ABSENT, AttendanceDayStatus.INCOMPLETE} or day.is_late or day.overtime_minutes:
                summary['rows'].append(day)
        current += timedelta(days=1)
    return summary


def list_org_attendance_days(organisation, *, target_date=None, employee_id=None, status_value=''):
    attendance_date = target_date or _policy_local_date(get_default_attendance_policy(organisation))
    employees = Employee.objects.filter(organisation=organisation, status=EmployeeStatus.ACTIVE)
    if employee_id:
        employees = employees.filter(id=employee_id)
    day_summaries = [recalculate_attendance_day(employee, attendance_date) for employee in employees.select_related('user', 'office_location')]
    if status_value:
        day_summaries = [item for item in day_summaries if item.status == status_value]
    return sorted(day_summaries, key=lambda item: (item.employee.employee_code or '', item.employee.user.full_name))


def list_attendance_regularizations_for_org(organisation, *, status_value=''):
    queryset = AttendanceRegularizationRequest.objects.filter(organisation=organisation).select_related(
        'employee__user',
        'attendance_day',
        'approval_run',
    )
    if status_value:
        queryset = queryset.filter(status=status_value)
    return queryset.order_by('-created_at')


def get_payroll_attendance_summary(employee, *, period_start, period_end):
    total_paid_fraction = ZERO
    total_overtime_minutes = 0
    current = period_start
    while current <= period_end:
        day = recalculate_attendance_day(employee, current)
        total_paid_fraction += _decimal(day.paid_fraction)
        total_overtime_minutes += day.overtime_minutes
        current += timedelta(days=1)
    total_days = Decimal((period_end - period_start).days + 1)
    lop_days = max(ZERO, total_days - total_paid_fraction)
    return {
        'paid_days': int(total_paid_fraction) if total_paid_fraction == total_paid_fraction.to_integral() else float(total_paid_fraction),
        'paid_fraction': _decimal(total_paid_fraction),
        'lop_days': _decimal(lop_days),
        'overtime_minutes': total_overtime_minutes,
    }


def import_attendance_sheet(*, organisation, uploaded_by, uploaded_file):
    _ensure_xlsx(uploaded_file)
    rows = _load_sheet_rows(uploaded_file, ATTENDANCE_HEADERS)
    if not rows:
        raise ValueError('The workbook does not contain any attendance rows.')

    policy = get_default_attendance_policy(organisation)
    with transaction.atomic():
        job = AttendanceImportJob.objects.create(
            organisation=organisation,
            uploaded_by=uploaded_by,
            mode=AttendanceImportMode.ATTENDANCE_SHEET,
            status=AttendanceImportStatus.FAILED,
            original_filename=getattr(uploaded_file, 'name', 'attendance.xlsx'),
            total_rows=len(rows),
        )

        seen_keys = set()
        valid_rows = []
        for row in rows:
            employee_code = _strip_string(row.get('employee_code')).upper()
            try:
                normalized_code, employee = _resolve_employee(organisation, employee_code)
                attendance_date = _parse_date(row.get('date'))
                check_in_at = _make_aware_datetime(attendance_date, _parse_time_value(row.get('check_in'), label='Check in'), policy)
                check_out_at = _make_aware_datetime(attendance_date, _parse_time_value(row.get('check_out'), label='Check out'), policy)
                if check_out_at < check_in_at:
                    raise ValueError('Check out cannot be earlier than check in.')
                dedupe_key = (employee.id, attendance_date)
                if dedupe_key in seen_keys:
                    raise ValueError('Only one attendance row per employee and date is allowed in the same upload.')
                seen_keys.add(dedupe_key)

                AttendanceImportRow.objects.create(
                    job=job,
                    row_number=row['row_number'],
                    employee_code=normalized_code,
                    employee=employee,
                    attendance_date=attendance_date,
                    check_in_at=check_in_at,
                    check_out_at=check_out_at,
                    status=AttendanceImportRowStatus.VALID,
                )
                valid_rows.append((employee, attendance_date, check_in_at, check_out_at))
            except ValueError as exc:
                AttendanceImportRow.objects.create(
                    job=job,
                    row_number=row['row_number'],
                    employee_code=employee_code,
                    status=AttendanceImportRowStatus.ERROR,
                    error_message=str(exc),
                )

        job.valid_rows = len(valid_rows)
        job.error_rows = job.rows.filter(status=AttendanceImportRowStatus.ERROR).count()
        if job.error_rows:
            job.status = AttendanceImportStatus.FAILED
            job.save(update_fields=['valid_rows', 'error_rows', 'status', 'modified_at'])
            log_audit_event(uploaded_by, 'attendance.import.failed', organisation=organisation, target=job, payload={'mode': job.mode})
            return job

        posted_rows = 0
        for employee, attendance_date, check_in_at, check_out_at in valid_rows:
            AttendanceRecord.objects.update_or_create(
                organisation=organisation,
                employee=employee,
                attendance_date=attendance_date,
                defaults={
                    'check_in_at': check_in_at,
                    'check_out_at': check_out_at,
                    'source': AttendanceRecordSource.EXCEL_IMPORT,
                    'import_job': job,
                },
            )
            recalculate_attendance_day(employee, attendance_date, actor=uploaded_by)
            posted_rows += 1

        job.rows.filter(status=AttendanceImportRowStatus.VALID).update(status=AttendanceImportRowStatus.POSTED)
        job.status = AttendanceImportStatus.POSTED
        job.posted_rows = posted_rows
        job.save(update_fields=['valid_rows', 'error_rows', 'status', 'posted_rows', 'modified_at'])
        log_audit_event(uploaded_by, 'attendance.import.posted', organisation=organisation, target=job, payload={'mode': job.mode, 'posted_rows': posted_rows})
        return job


def import_punch_sheet(*, organisation, uploaded_by, uploaded_file):
    _ensure_xlsx(uploaded_file)
    rows = _load_sheet_rows(uploaded_file, PUNCH_HEADERS)
    if not rows:
        raise ValueError('The workbook does not contain any punch rows.')

    policy = get_default_attendance_policy(organisation)
    with transaction.atomic():
        job = AttendanceImportJob.objects.create(
            organisation=organisation,
            uploaded_by=uploaded_by,
            mode=AttendanceImportMode.PUNCH_SHEET,
            status=AttendanceImportStatus.FAILED,
            original_filename=getattr(uploaded_file, 'name', 'attendance-punches.xlsx'),
            total_rows=len(rows),
        )
        grouped_punches = defaultdict(list)
        for row in rows:
            employee_code = _strip_string(row.get('employee_code')).upper()
            try:
                normalized_code, employee = _resolve_employee(organisation, employee_code)
                attendance_date = _parse_date(row.get('date'))
                punch_at = _make_aware_datetime(attendance_date, _parse_time_value(row.get('punch_time'), label='Punch time'), policy)
                grouped_punches[(employee.id, attendance_date)].append((normalized_code, employee, punch_at, row['row_number']))
            except ValueError as exc:
                AttendanceImportRow.objects.create(
                    job=job,
                    row_number=row['row_number'],
                    employee_code=employee_code,
                    status=AttendanceImportRowStatus.ERROR,
                    error_message=str(exc),
                )

        generated_rows = 0
        for (_employee_id, attendance_date), punch_rows in sorted(grouped_punches.items(), key=lambda item: (item[0][1], item[0][0])):
            punch_rows.sort(key=lambda item: item[2])
            employee_code, employee = punch_rows[0][0], punch_rows[0][1]
            check_in_at = punch_rows[0][2]
            check_out_at = punch_rows[-1][2] if len(punch_rows) > 1 else None
            status_value = AttendanceImportRowStatus.VALID if check_out_at else AttendanceImportRowStatus.INCOMPLETE
            error_message = '' if check_out_at else 'Only one punch was found for this employee on this date. Add a checkout time before re-uploading.'
            AttendanceImportRow.objects.create(
                job=job,
                row_number=generated_rows + 2,
                employee_code=employee_code,
                employee=employee,
                attendance_date=attendance_date,
                check_in_at=check_in_at,
                check_out_at=check_out_at,
                raw_punch_times=[timezone.localtime(item[2]).strftime('%H:%M:%S') for item in punch_rows],
                status=status_value,
                error_message=error_message,
                metadata={'source_row_numbers': [item[3] for item in punch_rows], 'punch_count': len(punch_rows)},
            )
            generated_rows += 1

        job.valid_rows = generated_rows
        job.error_rows = job.rows.filter(status=AttendanceImportRowStatus.ERROR).count()
        job.status = AttendanceImportStatus.READY_FOR_REVIEW if generated_rows else AttendanceImportStatus.FAILED
        job.save(update_fields=['valid_rows', 'error_rows', 'status', 'modified_at'])
        log_audit_event(uploaded_by, 'attendance.import.normalized', organisation=organisation, target=job, payload={'mode': job.mode, 'generated_rows': generated_rows})
        return job
